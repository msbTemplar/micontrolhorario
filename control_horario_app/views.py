from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
import pytz
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import UserCreationForm
from django.contrib import messages
from .models import TimeEntry, TimeEntryModificationRequest, User 
from django.utils import timezone
from django.core.mail import send_mail, BadHeaderError # Importar BadHeaderError
from django.template.loader import render_to_string
from django.utils.html import strip_tags
import logging # Importar el módulo de logging
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger # Importar Paginator
from django.db.models import Q # Importar Q para consultas complejas
from datetime import datetime, timedelta # Importar datetime y timedelta
from django.db.models import Sum, F # Asegúrate de que Sum y F estén aquí
# Nuevas importaciones para CSV
import csv
from django.http import HttpResponse

# Nuevas importaciones para Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import io # Para manejar el archivo en memoria

# Nuevas importaciones para PDF
from weasyprint import HTML, CSS # Importar HTML y CSS de WeasyPrint

from django.conf import settings # Para acceder a STATIC_ROOT si necesitas cargar CSS/Imágenes externas
import os # Para construir rutas de archivos
from .forms import TimeEntryEditForm, TimeEntryRequestModificationForm, UserProfileEditForm, PasswordChangingForm # NUEVAS IMPORTACIONES DE FORMULARIOS
# Importaciones de Django para la gestión de contraseñas
from django.contrib.auth import update_session_auth_hash # Necesario para mantener al usuario logeado después de cambiar contraseña
from django.contrib.auth.hashers import make_password # Añadir si decides permitir cambiar la contraseña desde el formulario de perfil (no recomendado)
from django.core.mail import EmailMultiAlternatives # <- Aquí se importa
from django.urls import reverse

# Importar el modelo User si no lo has hecho ya para is_admin
from django.contrib.auth import get_user_model
User = get_user_model()


# Configurar el logger
logger = logging.getLogger(__name__) # Obtiene un logger para este módulo

# Función para verificar si el usuario es admin
def is_admin(user):
    return user.is_staff

# Vista para listar las solicitudes de modificación (solo para admins)
@login_required
@user_passes_test(is_admin, login_url='/dashboard/') # Redirige si no es admin
def admin_modification_requests(request):
    pending_requests = TimeEntryModificationRequest.objects.filter(status='pending').order_by('requested_at')
    
    context = {
        'pending_requests': pending_requests
    }
    return render(request, 'control_horario_app/admin_requests_list.html', context)

# Vista para detalles y aprobación/rechazo de una solicitud
@login_required
@user_passes_test(is_admin, login_url='/dashboard/')
def review_modification_request(request, request_id):
    # Obtener el objeto de la solicitud de modificación
    modification_request = get_object_or_404(TimeEntryModificationRequest, id=request_id)

    if request.method == 'POST':
        action = request.POST.get('action')
        admin_notes = request.POST.get('admin_notes', '').strip()

        if modification_request.status != 'pending':
            messages.error(request, 'Esta solicitud ya ha sido procesada.')
            return redirect('admin_modification_requests')

        if action == 'approve':
            try:
                original_entry = modification_request.original_entry
                original_entry.start_time = modification_request.requested_start_time
                original_entry.end_time = modification_request.requested_end_time
                original_entry.break_start = modification_request.requested_break_start
                original_entry.break_end = modification_request.requested_break_end
                original_entry.is_edited = True
                original_entry.last_edited = timezone.now()
                original_entry.save()

                modification_request.status = 'approved'
                modification_request.admin_notes = admin_notes if admin_notes else 'Solicitud aprobada por el administrador.'
                modification_request.reviewer = request.user
                modification_request.reviewed_at = timezone.now()
                modification_request.save()

                messages.success(request, 'La solicitud ha sido aprobada y el registro de jornada actualizado.')
                send_request_status_email(modification_request, 'approved', request)
                
            except Exception as e:
                messages.error(request, f'Ocurrió un error al aprobar la solicitud: {e}')
                # Pasa el objeto correcto a la plantilla en caso de error
                return render(request, 'control_horario_app/admin_review_request.html', {'modification_request_obj': modification_request, 'error_message': f'Error al aprobar: {e}'})

        elif action == 'reject':
            if not admin_notes:
                messages.error(request, 'Debes proporcionar una razón para rechazar la solicitud.')
                # Pasa el objeto correcto a la plantilla en caso de error
                return render(request, 'control_horario_app/admin_review_request.html', {'modification_request_obj': modification_request, 'error_message': 'La razón de rechazo es obligatoria.'})

            modification_request.status = 'rejected'
            modification_request.admin_notes = admin_notes
            modification_request.reviewer = request.user
            modification_request.reviewed_at = timezone.now()
            modification_request.save()
            
            messages.info(request, 'La solicitud ha sido rechazada.')
            send_request_status_email(modification_request, 'rejected', request)

        return redirect('admin_modification_requests')

    # Si la solicitud es GET, simplemente renderiza la página
    context = {
        'modification_request_obj': modification_request, # <-- CAMBIADO AQUI
    }
    return render(request, 'control_horario_app/admin_review_request.html', context)



# Función auxiliar para enviar email al usuario que hizo la solicitud
def send_request_status_email(modification_request, status, request_obj):
    user_email = modification_request.requesting_user.email
    if not user_email:
        print(f"Advertencia: El usuario {modification_request.requesting_user.username} no tiene email para notificar el estado de la solicitud.")
        return

    subject = f"Tu Solicitud de Modificación de Jornada ha sido {status.capitalize()}"
    
    # URL de la jornada original para que el usuario pueda ver el cambio
    original_entry_url = request_obj.build_absolute_uri(reverse('edit_time_entry', args=[modification_request.original_entry.id]))

    context = {
        'modification_request': modification_request,
        'status': status,
        'original_entry': modification_request.original_entry,
        'original_entry_url': original_entry_url,
        'site_url_logo': 'https://i.imgur.com/oNrnCFF.jpeg', # Asegúrate de que esta URL es accesible
    }

    html_message = render_to_string(f'control_horario_app/user_request_status_email_{status.lower()}.html', context)
    plain_message = strip_tags(html_message)

    email = EmailMultiAlternatives(
        subject,
        plain_message,
        settings.DEFAULT_FROM_EMAIL,
        [user_email]
    )
    email.attach_alternative(html_message, "text/html")
    email.send(fail_silently=False)


# --- NUEVA VISTA: Perfil de Usuario ---
@login_required
def profile_view(request):
    if request.method == 'POST':
        form = UserProfileEditForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tu perfil ha sido actualizado exitosamente.')
            return redirect('profile') # Redirige de nuevo a la página de perfil
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Error en {field}: {error}")
    else:
        form = UserProfileEditForm(instance=request.user)
    
    context = {
        'form': form
    }
    return render(request, 'control_horario_app/profile.html', context)

# --- NUEVA VISTA: Cambiar Contraseña ---
@login_required
def change_password_view(request):
    if request.method == 'POST':
        form = PasswordChangingForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Importante: mantiene al usuario logeado
            messages.success(request, 'Tu contraseña ha sido cambiada exitosamente.')
            return redirect('profile') # Redirige al perfil o a donde quieras después del cambio
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Error en {field}: {error}")
    else:
        form = PasswordChangingForm(request.user)
    
    context = {
        'form': form
    }
    return render(request, 'control_horario_app/change_password.html', context)


@login_required
def edit_time_entry(request, entry_id):
    entry = get_object_or_404(TimeEntry, id=entry_id)

    # Permiso de acceso: Solo el propietario del registro O un administrador puede acceder a esta página.
    if not (request.user == entry.user or is_admin(request.user)):
        messages.error(request, 'No tienes permiso para editar este registro.')
        return redirect('dashboard')

    # Determinar qué formulario se usará y si es una edición directa o una solicitud.
    current_is_admin = is_admin(request.user)

    if request.method == 'POST':
        if current_is_admin: # Si es un admin, intentará editar directamente
            form = TimeEntryEditForm(request.POST, instance=entry)
            if form.is_valid():
                try:
                    # Guardar la instancia SIN commitear aún para poder añadir los campos de auditoría
                    updated_entry = form.save(commit=False)
                    updated_entry.is_edited = True
                    updated_entry.last_edited = timezone.now()
                    updated_entry.save() # Guarda la instancia en la base de datos AHORA

                    # Lógica para registrar la razón del admin en TimeEntryModificationRequest
                    admin_reason = form.cleaned_data.get('reason')
                    if admin_reason: # Solo crea la solicitud de modificación si el admin proporcionó una razón
                        TimeEntryModificationRequest.objects.create(
                            original_entry=updated_entry, # Usa la instancia que acaba de ser guardada
                            requesting_user=request.user,
                            requested_start_time=updated_entry.start_time,
                            requested_end_time=updated_entry.end_time,
                            requested_break_start=updated_entry.break_start,
                            requested_break_end=updated_entry.break_end,
                            reason=admin_reason,
                            status='approved',
                            reviewer=request.user,
                            reviewed_at=timezone.now(),
                            admin_notes=admin_reason
                        )
                    messages.success(request, 'El registro de jornada ha sido actualizado directamente por el administrador.')
                    return redirect('dashboard')
                except Exception as e:
                    # Captura cualquier error que ocurra durante el proceso de guardado y lo muestra
                    messages.error(request, f'Ocurrió un error al guardar el registro: {e}. Por favor, contacte al soporte.')
                    print(f"ERROR al guardar registro (admin): {e}") # Imprime el error en la consola del servidor para depuración
            else:
                # Si el formulario no es válido, se mostrarán los errores en la plantilla
                print(f"Errores en el formulario de edición (admin): {form.errors}") # También imprime los errores en la consola
                messages.error(request, 'Por favor, corrige los errores en el formulario de edición directa.')
                # El formulario con errores se pasará al contexto para que se muestren en la plantilla
        else: # Si NO es un admin, es una solicitud de cambio (este bloque se mantiene igual que antes)
            form = TimeEntryRequestModificationForm(request.POST, original_entry=entry)
            if form.is_valid():
                try:
                    modification_request = TimeEntryModificationRequest.objects.create(
                        original_entry=entry,
                        requesting_user=request.user,
                        requested_start_time=form.cleaned_data.get('combined_start_time'),
                        requested_end_time=form.cleaned_data.get('combined_end_time'),
                        requested_break_start=form.cleaned_data.get('combined_break_start'),
                        requested_break_end=form.cleaned_data.get('combined_break_end'),
                        reason=form.cleaned_data.get('reason', ''),
                        status='pending'
                    )
                    messages.success(request, 'Su solicitud de modificación ha sido enviada al administrador y está pendiente de revisión.')

                    # Lógica de envío de correo
                    admin_users = User.objects.filter(is_staff=True, is_active=True)
                    admin_emails = [u.email for u in admin_users if u.email]
                    if admin_emails:
                        subject = f"Nueva Solicitud de Modificación de Jornada de {request.user.username}"
                        context_email = {
                            'requesting_user': request.user,
                            'original_entry': entry,
                            'modification_request': modification_request,
                            'requested_start_time': modification_request.requested_start_time,
                            'requested_end_time': modification_request.requested_end_time,
                            'requested_break_start': modification_request.requested_break_start,
                            'requested_break_end': modification_request.requested_break_end,
                            'reason': modification_request.reason,
                            'request_url': request.build_absolute_uri(reverse('admin:%s_%s_change' % (modification_request._meta.app_label, modification_request._meta.model_name), args=[modification_request.id])),
                            'admin_name': 'Administrador',
                            'company_name': 'Tu Empresa',
                            'site_url_logo': 'https://i.imgur.com/oNrnCFF.jpeg'
                        }
                        html_message = render_to_string('control_horario_app/admin_request_email.html', context_email)
                        plain_message = strip_tags(html_message)
                        email = EmailMultiAlternatives(
                            subject, plain_message, settings.DEFAULT_FROM_EMAIL, admin_emails
                        )
                        email.attach_alternative(html_message, "text/html")
                        email.send(fail_silently=False)
                    else:
                        messages.warning(request, 'Su solicitud de cambio ha sido guardada, pero no hay administradores configurados para recibir el correo de notificación.')
                    return redirect('dashboard')
                except Exception as e:
                    messages.error(request, f'Ocurrió un error al procesar su solicitud de cambio: {e}. Por favor, contacte al soporte.')
                    print(f"ERROR al crear solicitud o enviar email: {e}")
            else:
                print(f"Errores en el formulario de solicitud (usuario): {form.errors}")
                messages.error(request, 'Por favor, corrige los errores en el formulario de solicitud de cambio.')

    else: # GET request (cuando se carga la página por primera vez)
        # Inicializar el formulario apropiado con los datos de la instancia existente.
        if current_is_admin:
            form = TimeEntryEditForm(instance=entry)
        else:
            form = TimeEntryRequestModificationForm(original_entry=entry)

    # El formulario (ya sea con datos precargados o con errores POST) se pasa al contexto
    context = {
        'form': form, # Ahora solo pasamos una variable 'form' a la plantilla
        'entry': entry,
        'is_admin': current_is_admin,
    }
    return render(request, 'control_horario_app/edit_time_entry.html', context)





@login_required
def edit_time_entry2(request, entry_id):
    entry = get_object_or_404(TimeEntry, id=entry_id, user=request.user)
    
    is_admin = request.user.is_staff # O request.user.is_admin si usas un campo personalizado

    if request.method == 'POST':
        form = TimeEntryEditForm(request.POST) # No pases 'instance=entry' aquí para el POST, ya que no guardaremos directamente en el original TimeEntry para no-admins. Pasamos los datos del POST.
        
        if form.is_valid():
            # Obtener los datos limpios del formulario (estos son los valores PROPUESTOS)
            requested_start_time = form.cleaned_data['start_time']
            requested_end_time = form.cleaned_data.get('end_time')
            requested_break_start = form.cleaned_data.get('break_start')
            requested_break_end = form.cleaned_data.get('break_end')
            reason = form.cleaned_data.get('reason')

            if is_admin:
                # Si es administrador, aplicar los cambios directamente al registro original
                entry.start_time = requested_start_time
                entry.end_time = requested_end_time
                entry.break_start = requested_break_start
                entry.break_end = requested_break_end
                entry.notes = form.cleaned_data.get('notes', entry.notes) # Asegúrate de que las notas también se actualicen si el formulario las maneja. Si no tienes 'notes' en el form, elimina esta línea.
                
                # Actualizar campos de auditoría si existen
                entry.is_edited = True 
                entry.last_edited = timezone.now()
                
                entry.save()
                messages.success(request, 'Registro de jornada actualizado exitosamente.')
                return redirect('dashboard')
            else:
                # Si NO es administrador, crear una solicitud de modificación
                # y enviar un correo al administrador. NO MODIFICAMOS el TimeEntry original AQUÍ.
                
                # VALIDACIÓN ADICIONAL PARA NO-ADMINS: LA RAZÓN ES OBLIGATORIA
                if not reason or not reason.strip():
                    form.add_error('reason', 'Es obligatorio proporcionar una razón para la solicitud de cambio.')
                    messages.error(request, 'Por favor, corrige los errores en el formulario.')
                    # Vuelve a renderizar el formulario con los datos originales y el error.
                    # Necesitamos pasar un form con los datos que el usuario intentó enviar
                    # para que los campos no se reinicien.
                    return render(request, 'control_horario_app/edit_time_entry.html', {'form': form, 'entry': entry, 'is_admin': is_admin})


                try:
                    # Crear la solicitud en la base de datos
                    modification_request = TimeEntryModificationRequest.objects.create(
                        original_entry=entry, # Referencia al TimeEntry original
                        requesting_user=request.user,
                        requested_start_time=requested_start_time,
                        requested_end_time=requested_end_time,
                        requested_break_start=requested_break_start,
                        requested_break_end=requested_break_end,
                        reason=reason
                    )

                    # Enviar correo al administrador(es)
                    admin_users = User.objects.filter(is_staff=True, is_active=True) 
                    admin_emails = [u.email for u in admin_users if u.email]

                    if admin_emails:
                        subject = f"Nueva Solicitud de Modificación de Jornada de {request.user.username}"
                        
                        # URL para que el admin pueda revisar la solicitud
                        # Si tienes una vista específica para manejar solicitudes:
                        # request_url = request.build_absolute_uri(reverse('review_modification_request_detail', args=[modification_request.id]))
                        # Por ahora, podemos enviar al dashboard o a la página de edición del TimeEntry original
                        request_url = request.build_absolute_uri(f'/editar-jornada/{entry.id}/') # Envía al admin a la página de edición del TimeEntry original
                        
                        context = {
                            'requesting_user': request.user,
                            'original_entry': entry, # El registro actual
                            'modification_request': modification_request, # La solicitud recién creada
                            'requested_start_time': requested_start_time, # Pasar los datos solicitados individualmente para la plantilla
                            'requested_end_time': requested_end_time,
                            'requested_break_start': requested_break_start,
                            'requested_break_end': requested_break_end,
                            'reason': reason,
                            'request_url': request_url,
                            'admin_name': 'Administrador',
                            'company_name': 'Tu Empresa', 
                            'site_url_logo': 'https://i.imgur.com/oNrnCFF.jpeg' # La URL del logo que ya te funciona
                        }
                        
                        html_message = render_to_string('control_horario_app/admin_request_email.html', context)
                        plain_message = strip_tags(html_message)

                        email = EmailMultiAlternatives(
                            subject,
                            plain_message,
                            settings.DEFAULT_FROM_EMAIL,
                            admin_emails
                        )
                        email.attach_alternative(html_message, "text/html")
                        
                        email.send(fail_silently=False)
                        messages.success(request, 'Su solicitud de cambio ha sido enviada al administrador y está pendiente de revisión.')
                    else:
                        messages.warning(request, 'Su solicitud de cambio ha sido guardada, pero no hay administradores configurados para recibir el correo de notificación.')
                    
                    return redirect('dashboard')

                except Exception as e:
                    messages.error(request, f'Ocurrió un error al procesar su solicitud de cambio: {e}. Por favor, contacte al soporte.')
                    # Puedes loguear el error aquí
                    print(f"ERROR al crear solicitud o enviar email: {e}")
                    return render(request, 'control_horario_app/edit_time_entry.html', {'form': form, 'entry': entry, 'is_admin': is_admin})

        else:
            # Si el formulario no es válido (ej. errores de fecha/hora), mostrar errores
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
            # Importante: para que el formulario se muestre con los datos y errores
            # Correctamente, si falla en la validación, pasamos el `form` con `request.POST`
            # que ya contiene los datos y los errores.
            # No recrees el formulario con `instance=entry` si viene de un POST inválido.
            return render(request, 'control_horario_app/edit_time_entry.html', {'form': form, 'entry': entry, 'is_admin': is_admin})
    else:
        # Método GET: inicializar el formulario con los datos existentes del TimeEntry
        initial_data = {
            'start_date': timezone.localtime(entry.start_time).date(),
            'start_time_part': timezone.localtime(entry.start_time).time(),
        }
        if entry.end_time:
            local_end_time = timezone.localtime(entry.end_time)
            initial_data['end_date'] = local_end_time.date()
            initial_data['end_time_part'] = local_end_time.time()
        if entry.break_start:
            local_break_start = timezone.localtime(entry.break_start)
            initial_data['break_start_date'] = local_break_start.date()
            initial_data['break_start_time_part'] = local_break_start.time()
        if entry.break_end:
            local_break_end = timezone.localtime(entry.break_end)
            initial_data['break_end_date'] = local_break_end.date()
            initial_data['break_end_time_part'] = local_break_end.time()
            
        form = TimeEntryEditForm(initial=initial_data) # Para GET, solo initial, sin instance
    
    return render(request, 'control_horario_app/edit_time_entry.html', {'form': form, 'entry': entry, 'is_admin': is_admin})





# --- NUEVA VISTA: Editar Registro de Jornada ---
@login_required
def edit_time_entry1(request, entry_id):
    entry = get_object_or_404(TimeEntry, id=entry_id, user=request.user)
    spain_tz = pytz.timezone('Europe/Madrid')

    # No permitir editar una jornada que está en curso (end_time es nulo)
    if not entry.end_time:
        messages.error(request, "No puedes editar una jornada que aún está en curso.")
        return redirect('dashboard')

    if request.method == 'POST':
        form = TimeEntryEditForm(request.POST, instance=entry)
        if form.is_valid():
            # El formulario ya combina y maneja las zonas horarias en su método clean
            # y las validaciones lógicas de tiempo.
            form.save()
            # Si quieres registrar la edición
            # entry.is_edited = True
            # entry.last_edited = timezone.now().astimezone(spain_tz)
            # entry.save()

            messages.success(request, "Registro de jornada actualizado exitosamente.")
            return redirect('dashboard')
        else:
            # Pasa los errores del formulario a los mensajes para el usuario
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Error en {field}: {error}")
    else:
        # Si es una solicitud GET, inicializa el formulario con los datos existentes
        form = TimeEntryEditForm(instance=entry)
    
    context = {
        'form': form,
        'entry': entry, # Pasar la entrada para mostrar detalles en la plantilla
    }
    return render(request, 'control_horario_app/edit_time_entry.html', context)

# --- NUEVA VISTA: Eliminar Registro de Jornada ---
@login_required
def delete_time_entry(request, entry_id):
    entry = get_object_or_404(TimeEntry, id=entry_id, user=request.user)

    # Solo permitir eliminar si la jornada está finalizada
    if not entry.end_time:
        messages.error(request, "No puedes eliminar una jornada que aún está en curso.")
        return redirect('dashboard')

    if request.method == 'POST':
        entry.delete()
        messages.success(request, "Registro de jornada eliminado exitosamente.")
    else:
        messages.error(request, "Método no permitido para la eliminación directa. Usa el formulario de confirmación.")
    
    return redirect('dashboard')


@login_required
def export_timesheet_pdf(request):
    spain_tz = pytz.timezone('Europe/Madrid')

    #entries = TimeEntry.objects.filter(user=request.user).order_by('-start_time')
    
    # Determinar si el usuario es un administrador
    # NOTA: Asegúrate de que el nombre del grupo 'Administradores' sea EXACTO
    is_admin = request.user.groups.filter(name='Administradores').exists()

    # Si es administrador, recupera todos los registros. Si no, solo los del usuario actual.
    if is_admin:
        entries = TimeEntry.objects.all()
    else:
        entries = TimeEntry.objects.filter(user=request.user)

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    if start_date_str:
        try:
            start_date_naive = datetime.strptime(start_date_str, '%Y-%m-%d')
            start_date = timezone.make_aware(datetime.combine(start_date_naive, datetime.min.time()), spain_tz)
            entries = entries.filter(start_time__gte=start_date)
        except ValueError:
            pass

    if end_date_str:
        try:
            end_date_naive = datetime.strptime(end_date_str, '%Y-%m-%d')
            end_date = timezone.make_aware(datetime.combine(end_date_naive, datetime.max.time()), spain_tz)
            entries = entries.filter(start_time__lte=end_date)
        except ValueError:
            pass

    # Generar el nombre del archivo con fecha y hora
    current_datetime = timezone.now().astimezone(spain_tz)
    timestamp_str = current_datetime.strftime('%Y%m%d_%H%M%S')
    file_name = f"registros_horarios_{request.user.username}_{timestamp_str}.pdf"

    # Preparar el contexto para la plantilla HTML del PDF
    context = {
        'user': request.user,
        'entries': entries, # Pasar todas las entradas filtradas
        'start_date': start_date_str,
        'end_date': end_date_str,
        'generated_date': current_datetime.strftime('%d/%m/%Y %H:%M:%S'),
        'spain_tz': spain_tz # Pasar la zona horaria para formato en plantilla si es necesario
    }
    
    # Renderizar la plantilla HTML para el PDF
    # Podrías crear una plantilla específica para PDF (ej. 'email/jornada_finalizada_pdf.html')
    # pero para empezar, podemos usar una versión adaptada de la de email.
    html_string = render_to_string('control_horario_app/email/registros_pdf.html', context) # NUEVA PLANTILLA HTML para PDF

    # Convertir el HTML a PDF usando WeasyPrint
    pdf_file = HTML(string=html_string).write_pdf()

    # Crear la respuesta HTTP para la descarga de PDF
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'
    return response



@login_required
def export_timesheet_xlsx(request):
    spain_tz = pytz.timezone('Europe/Madrid')

    #entries = TimeEntry.objects.filter(user=request.user).order_by('-start_time')
    
    # Determinar si el usuario es un administrador
    # NOTA: Asegúrate de que el nombre del grupo 'Administradores' sea EXACTO
    is_admin = request.user.groups.filter(name='Administradores').exists()

    # Si es administrador, recupera todos los registros. Si no, solo los del usuario actual.
    if is_admin:
        entries = TimeEntry.objects.all()
    else:
        entries = TimeEntry.objects.filter(user=request.user)

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    if start_date_str:
        try:
            start_date_naive = datetime.strptime(start_date_str, '%Y-%m-%d')
            start_date = timezone.make_aware(datetime.combine(start_date_naive, datetime.min.time()), spain_tz)
            entries = entries.filter(start_time__gte=start_date)
        except ValueError:
            pass

    if end_date_str:
        try:
            end_date_naive = datetime.strptime(end_date_str, '%Y-%m-%d')
            end_date = timezone.make_aware(datetime.combine(end_date_naive, datetime.max.time()), spain_tz)
            entries = entries.filter(start_time__lte=end_date)
        except ValueError:
            pass

    # Crear un nuevo libro de trabajo de Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Registros Horarios"

    # --- Estilos para el Excel ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid") # Verde
    # Puedes cambiar el color a tu gusto, por ejemplo:
    # header_fill = PatternFill(start_color="0056b3", end_color="0056b3", fill_type="solid") # Azul
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    center_aligned_text = Alignment(horizontal="center", vertical="center")

    # Definir la fila de encabezado
    headers = [
        'ID Registro', 'Usuario', 'Fecha Inicio', 'Hora Inicio', 'Fecha Fin', 'Hora Fin',
        'Fecha Inicio Descanso', 'Hora Inicio Descanso', 'Fecha Fin Descanso', 'Hora Fin Descanso',
        'Duracion Neta Horas'
    ]
    sheet.append(headers)

    # Aplicar estilos al encabezado
    for col_num, cell in enumerate(sheet[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned_text
        cell.border = thin_border
        sheet.column_dimensions[get_column_letter(col_num)].width = 18 # Ajustar ancho de columna

    # Escribir los datos de cada registro
    for entry in entries:
        row_data = [
            entry.id,
            entry.user.username,
            entry.start_time.astimezone(spain_tz).strftime('%Y-%m-%d'),
            entry.start_time.astimezone(spain_tz).strftime('%H:%M:%S'),
            entry.end_time.astimezone(spain_tz).strftime('%Y-%m-%d') if entry.end_time else '',
            entry.end_time.astimezone(spain_tz).strftime('%H:%M:%S') if entry.end_time else '',
            entry.break_start.astimezone(spain_tz).strftime('%Y-%m-%d') if entry.break_start else '',
            entry.break_start.astimezone(spain_tz).strftime('%H:%M:%S') if entry.break_start else '',
            entry.break_end.astimezone(spain_tz).strftime('%Y-%m-%d') if entry.break_end else '',
            entry.break_end.astimezone(spain_tz).strftime('%H:%M:%S') if entry.break_end else '',
            f"{entry.duration:.2f}" if entry.duration is not None else 'N/A'
        ]
        sheet.append(row_data)
        # Aplicar borde a las celdas de datos
        for cell in sheet[sheet.max_row]:
            cell.border = thin_border

    # Generar el nombre del archivo con fecha y hora
    current_datetime = timezone.now().astimezone(spain_tz)
    timestamp_str = current_datetime.strftime('%Y%m%d_%H%M%S')
    file_name = f"registros_horarios_{request.user.username}_{timestamp_str}.xlsx"

    # Guardar el libro de trabajo en un objeto BytesIO (en memoria)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0) # Mover el puntero al inicio del archivo

    # Crear la respuesta HTTP para la descarga de Excel
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'
    return response

@login_required
def export_timesheet_csv(request):
    # Obtener la zona horaria de España para consistencia
    spain_tz = pytz.timezone('Europe/Madrid')

    # Filtrar registros por el usuario actual
    #entries = TimeEntry.objects.filter(user=request.user).order_by('-start_time')
    
    # Determinar si el usuario es un administrador
    # NOTA: Asegúrate de que el nombre del grupo 'Administradores' sea EXACTO
    is_admin = request.user.groups.filter(name='Administradores').exists()

    # Si es administrador, recupera todos los registros. Si no, solo los del usuario actual.
    if is_admin:
        entries = TimeEntry.objects.all()
    else:
        entries = TimeEntry.objects.filter(user=request.user)

    # Obtener fechas de filtro si están presentes en la URL (opcional, pero buena práctica)
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    if start_date_str:
        try:
            start_date_naive = datetime.strptime(start_date_str, '%Y-%m-%d')
            start_date = timezone.make_aware(datetime.combine(start_date_naive, datetime.min.time()), spain_tz)
            entries = entries.filter(start_time__gte=start_date)
        except ValueError:
            pass

    if end_date_str:
        try:
            end_date_naive = datetime.strptime(end_date_str, '%Y-%m-%d')
            end_date = timezone.make_aware(datetime.combine(end_date_naive, datetime.max.time()), spain_tz)
            entries = entries.filter(start_time__lte=end_date)
        except ValueError:
            pass
    # --- NUEVA LÓGICA PARA EL NOMBRE DEL ARCHIVO CON FECHA Y HORA ---
    current_datetime = timezone.now().astimezone(spain_tz)
    timestamp_str = current_datetime.strftime('%Y%m%d_%H%M%S') # Formato: AAAA-MM-DD_HHMMSS
    file_name = f"registros_horarios_{request.user.username}_{timestamp_str}.csv"
    # --- FIN NUEVA LÓGICA ---
    
    # Crear la respuesta HTTP con el tipo de contenido CSV
    response = HttpResponse(content_type='text/csv')
    # Definir el nombre del archivo que se descargará
    # response['Content-Disposition'] = f'attachment; filename="registros_horarios_{request.user.username}.csv"'
    response['Content-Disposition'] = f'attachment; filename="{file_name}"' # Usa la nueva variable file_name
    
    
    writer = csv.writer(response)

    # Escribir la fila de encabezado
    writer.writerow([
        'ID Registro',
        'Usuario',
        'Fecha Inicio',
        'Hora Inicio',
        'Fecha Fin',
        'Hora Fin',
        'Fecha Inicio Descanso',
        'Hora Inicio Descanso',
        'Fecha Fin Descanso',
        'Hora Fin Descanso',
        'Duracion Neta Horas'
    ])

    # Escribir los datos de cada registro
    for entry in entries:
        writer.writerow([
            entry.id,
            entry.user.username,
            entry.start_time.astimezone(spain_tz).strftime('%Y-%m-%d'), # Fecha en zona horaria local
            entry.start_time.astimezone(spain_tz).strftime('%H:%M:%S'), # Hora en zona horaria local
            entry.end_time.astimezone(spain_tz).strftime('%Y-%m-%d') if entry.end_time else '',
            entry.end_time.astimezone(spain_tz).strftime('%H:%M:%S') if entry.end_time else '',
            entry.break_start.astimezone(spain_tz).strftime('%Y-%m-%d') if entry.break_start else '',
            entry.break_start.astimezone(spain_tz).strftime('%H:%M:%S') if entry.break_start else '',
            entry.break_end.astimezone(spain_tz).strftime('%Y-%m-%d') if entry.break_end else '',
            entry.break_end.astimezone(spain_tz).strftime('%H:%M:%S') if entry.break_end else '',
            f"{entry.duration:.2f}" if entry.duration is not None else 'N/A' # Formato para la duración neta
        ])

    return response

def custom_logout_view(request):
    logout(request)
    return redirect('login') 

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')
    return render(request, 'control_horario_app/login.html')

def register_view(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cuenta creada exitosamente')
            return redirect('login')
    else:
        form = UserCreationForm()
    return render(request, 'control_horario_app/register.html', {'form': form})


@login_required
def dashboard_view(request):
    spain_tz = pytz.timezone('Europe/Madrid')
    
    # Obtener la fecha y hora actuales en la zona horaria de España
    now_spain = timezone.now().astimezone(spain_tz)
    today_spain = now_spain.date()

    # --- Lógica de Filtro por Fechas (la que ya tienes) ---
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    #entries = TimeEntry.objects.filter(user=request.user)
    
    # Determinar si el usuario es un administrador
    # NOTA: Asegúrate de que el nombre del grupo 'Administradores' sea EXACTO
    is_admin = request.user.groups.filter(name='Administradores').exists()

    # Si es administrador, recupera todos los registros. Si no, solo los del usuario actual.
    if is_admin:
        entries = TimeEntry.objects.all()
    else:
        entries = TimeEntry.objects.filter(user=request.user)

    if start_date_str:
        try:
            start_date_naive = datetime.strptime(start_date_str, '%Y-%m-%d')
            start_date = timezone.make_aware(datetime.combine(start_date_naive, datetime.min.time()), spain_tz)
            entries = entries.filter(start_time__gte=start_date)
        except ValueError:
            pass

    if end_date_str:
        try:
            end_date_naive = datetime.strptime(end_date_str, '%Y-%m-%d')
            end_date = timezone.make_aware(datetime.combine(end_date_naive, datetime.max.time()), spain_tz)
            entries = entries.filter(start_time__lte=end_date)
        except ValueError:
            pass

    entries = entries.order_by('-start_time')

    # --- Lógica de Estadísticas Semanales y Mensuales ---

    # Estadísticas Semanales
    # Calcular el inicio de la semana (Lunes) y fin de la semana (Domingo)
    # today_spain.weekday() devuelve 0 para Lunes, 6 para Domingo
    start_of_week = today_spain - timedelta(days=today_spain.weekday())
    end_of_week = start_of_week + timedelta(days=6)
    
    # Asegurarse de que sean objetos aware de la zona horaria y cubran todo el día
    start_of_week_aware = timezone.make_aware(datetime.combine(start_of_week, datetime.min.time()), spain_tz)
    end_of_week_aware = timezone.make_aware(datetime.combine(end_of_week, datetime.max.time()), spain_tz)

    weekly_entries = TimeEntry.objects.filter(
        user=request.user,
        start_time__gte=start_of_week_aware,
        start_time__lte=end_of_week_aware,
        end_time__isnull=False # Solo considerar jornadas finalizadas para el cálculo
    )

    total_weekly_seconds = 0
    for entry in weekly_entries:
        if entry.start_time and entry.end_time:
            work_duration = (entry.end_time - entry.start_time).total_seconds()
            break_duration = 0
            if entry.break_start and entry.break_end:
                break_duration = (entry.break_end - entry.break_start).total_seconds()
            total_weekly_seconds += (work_duration - break_duration)

    weekly_hours = total_weekly_seconds / 3600 if total_weekly_seconds else 0

    # Estadísticas Mensuales
    # Calcular el inicio del mes y fin del mes
    start_of_month = today_spain.replace(day=1)
    # Calcular el último día del mes: primer día del siguiente mes - 1 día
    if today_spain.month == 12:
        end_of_month = today_spain.replace(year=today_spain.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        end_of_month = today_spain.replace(month=today_spain.month + 1, day=1) - timedelta(days=1)

    start_of_month_aware = timezone.make_aware(datetime.combine(start_of_month, datetime.min.time()), spain_tz)
    end_of_month_aware = timezone.make_aware(datetime.combine(end_of_month, datetime.max.time()), spain_tz)

    monthly_entries = TimeEntry.objects.filter(
        user=request.user,
        start_time__gte=start_of_month_aware,
        start_time__lte=end_of_month_aware,
        end_time__isnull=False # Solo considerar jornadas finalizadas
    )

    total_monthly_seconds = 0
    for entry in monthly_entries:
        if entry.start_time and entry.end_time:
            work_duration = (entry.end_time - entry.start_time).total_seconds()
            break_duration = 0
            if entry.break_start and entry.break_end:
                break_duration = (entry.break_end - entry.break_start).total_seconds()
            total_monthly_seconds += (work_duration - break_duration)
            
    monthly_hours = total_monthly_seconds / 3600 if total_monthly_seconds else 0


    # --- Paginación (la que ya tienes) ---
    page = request.GET.get('page', 1)
    paginator = Paginator(entries, 10)

    try:
        entries_page = paginator.page(page)
    except PageNotAnInteger:
        entries_page = paginator.page(1)
    except EmptyPage:
        entries_page = paginator.page(paginator.num_pages)

    context = {
        'is_admin': is_admin,
        'entries': entries_page,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'weekly_hours': weekly_hours,
        'monthly_hours': monthly_hours,
        # AÑADE ESTAS LÍNEAS:
        'start_of_week': start_of_week,
        'end_of_week': end_of_week,
        'start_of_month': start_of_month,
        # Opcional: para mostrar el nombre del mes
        'month_name': start_of_month.strftime('%B %Y')
    }
    return render(request, 'control_horario_app/dashboard.html', context)


@login_required
def dashboard_view2(request):
    # Obtener la zona horaria de España (para consistencia con timesheet_view)
    spain_tz = pytz.timezone('Europe/Madrid')

    # --- Lógica de Filtrado por Fechas ---
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    entries = TimeEntry.objects.filter(user=request.user)

    # Si hay fechas, aplicar el filtro
    if start_date_str:
        try:
            # Convertir la fecha de inicio a un objeto datetime en la zona horaria correcta
            start_date_naive = datetime.strptime(start_date_str, '%Y-%m-%d')
            start_date = timezone.make_aware(datetime.combine(start_date_naive, datetime.min.time()), spain_tz)
            entries = entries.filter(start_time__gte=start_date)
        except ValueError:
            # Manejar el error si la fecha no es válida
            pass # Puedes añadir un message.error aquí si quieres

    if end_date_str:
        try:
            # Convertir la fecha de fin a un objeto datetime en la zona horaria correcta
            # Incluir el final del día (23:59:59)
            end_date_naive = datetime.strptime(end_date_str, '%Y-%m-%d')
            end_date = timezone.make_aware(datetime.combine(end_date_naive, datetime.max.time()), spain_tz)
            entries = entries.filter(start_time__lte=end_date)
        except ValueError:
            # Manejar el error si la fecha no es válida
            pass # Puedes añadir un message.error aquí si quieres

    entries = entries.order_by('-start_time') # Ordenar por fecha de inicio descendente

    # --- Lógica de Paginación ---
    page = request.GET.get('page', 1) # Obtener el número de página de la URL, por defecto 1
    paginator = Paginator(entries, 10) # Mostrar 10 registros por página

    try:
        entries_page = paginator.page(page)
    except PageNotAnInteger:
        # Si el parámetro 'page' no es un entero, entrega la primera página.
        entries_page = paginator.page(1)
    except EmptyPage:
        # Si la página está fuera de rango (ej. 9999), entrega la última página de resultados.
        entries_page = paginator.page(paginator.num_pages)

    context = {
        'entries': entries_page,
        'start_date': start_date_str, # Pasar las fechas de filtro para que se mantengan en el formulario
        'end_date': end_date_str,
    }
    return render(request, 'control_horario_app/dashboard.html', context)

@login_required
def dashboard_view1(request):
    entries = TimeEntry.objects.filter(user=request.user).order_by('-start_time')
    return render(request, 'control_horario_app/dashboard.html', {'entries': entries})

@login_required
def timesheet_view(request):
    spain_tz = pytz.timezone('Europe/Madrid')
    today_spain = timezone.now().astimezone(spain_tz).date()

    current_active_entry = TimeEntry.objects.filter(
        user=request.user,
        start_time__date=today_spain,
        end_time__isnull=True
    ).first()

    current_completed_entry = TimeEntry.objects.filter(
        user=request.user,
        start_time__date=today_spain,
        end_time__isnull=False
    ).first()
    
    show_completed_message = False

    if request.method == 'POST':
        action = request.POST.get('action')
        entry_id = request.POST.get('entry_id')

        if action == 'start':
            if not current_active_entry and not current_completed_entry:
                TimeEntry.objects.create(
                    user=request.user,
                    start_time=timezone.now()
                )
                messages.success(request, 'Jornada iniciada con éxito.')
            else:
                messages.warning(request, 'Ya tienes una jornada en curso o completada para hoy.')
            return redirect('timesheet') 
        
        if entry_id and current_active_entry: 
            try:
                entry = TimeEntry.objects.get(id=entry_id, user=request.user)
                if entry != current_active_entry:
                    messages.error(request, 'Error: La entrada que intentas modificar no es la actual.')
                    return redirect('timesheet') 

                if action == 'end':
                    entry.end_time = timezone.now()
                    entry.save()
                    
                    # --- Lógica de envío de email al finalizar la jornada ---
                    if entry.user.email: 
                        try:
                            subject = 'Resumen de tu Jornada Laboral Finalizada'
                            context = {
                                'user': request.user,
                                'entry': entry,
                            }
                            html_message = render_to_string('control_horario_app/email/jornada_finalizada.html', context)
                            plain_message = strip_tags(html_message)

                            send_mail(
                                subject,
                                plain_message,
                                None, 
                                [entry.user.email,'msb.caixa@gmail.com','msb.tesla@gmail.com', 'msb.coin@gmail.com', 'msb.duck@gmail.com', 'msebti2@gmail.com', 'papioles@gmail.com', 'msb.motive@gmail.com', 'msb.acer@gmail.com'], 
                                html_message=html_message,
                                fail_silently=False, 
                            )
                            messages.success(request, 'Jornada finalizada. Resumen enviado a tu correo.')
                        except BadHeaderError:
                            messages.error(request, 'Error: Encabezado de email inválido.')
                            logger.error(f"BadHeaderError al enviar email para usuario {request.user.username}")
                        except Exception as e:
                            messages.error(request, 'Error al enviar el resumen de jornada por correo.')
                            logger.error(f"Error al enviar email para usuario {request.user.username}: {e}")
                    else:
                        messages.warning(request, 'Jornada finalizada. No se pudo enviar el correo porque no tienes un email registrado.')

                    show_completed_message = True 

                elif action == 'break_start':
                    if not entry.break_start: 
                        entry.break_start = timezone.now()
                        entry.save()
                        messages.info(request, 'Descanso iniciado.')
                    else:
                        messages.warning(request, 'Ya habías iniciado un descanso.')

                elif action == 'break_end':
                    if entry.break_start and not entry.break_end: 
                        entry.break_end = timezone.now()
                        entry.save()
                        messages.info(request, 'Descanso finalizado.')
                    else:
                        messages.warning(request, 'No hay un descanso activo para finalizar.')
                
            except TimeEntry.DoesNotExist:
                messages.error(request, 'Error: La entrada de jornada no fue encontrada.')
                return redirect('dashboard') # Redirige a dashboard si la entrada no existe
            except Exception as e:
                messages.error(request, 'Ocurrió un error inesperado al procesar tu jornada.')
                logger.error(f"Error inesperado en timesheet_view para usuario {request.user.username}: {e}")
            
            return redirect('timesheet') 
    
    if current_completed_entry:
        show_completed_message = True

    return render(request, 'control_horario_app/timesheet.html', {
        'current_entry': current_active_entry,
        'show_completed_message': show_completed_message,
        'has_active_or_completed_entry_today': current_active_entry or current_completed_entry
    })


@login_required
def timesheet_view1(request):
    spain_tz = pytz.timezone('Europe/Madrid')
    today_spain = timezone.now().astimezone(spain_tz).date()

    # Primero, busca una jornada ABIERTA para el día de hoy.
    # Una jornada está abierta si start_time es hoy y end_time es NULL.
    current_active_entry = TimeEntry.objects.filter(
        user=request.user,
        start_time__date=today_spain,
        end_time__isnull=True
    ).first()

    # Luego, busca si ya hay una jornada CERRADA para el día de hoy.
    # Esto nos indicará si ya se completó una jornada hoy.
    current_completed_entry = TimeEntry.objects.filter(
        user=request.user,
        start_time__date=today_spain,
        end_time__isnull=False
    ).first()
    
    # Variable para controlar el mensaje en la plantilla
    show_completed_message = False

    if request.method == 'POST':
        action = request.POST.get('action')
        entry_id = request.POST.get('entry_id')

        if action == 'start':
            # Solo permitir iniciar una jornada si NO hay una jornada activa Y NO hay una jornada completada hoy
            if not current_active_entry and not current_completed_entry:
                TimeEntry.objects.create(
                    user=request.user,
                    start_time=timezone.now()
                )
            # Siempre redirigir a la misma vista para actualizar el estado de los botones
            return redirect('timesheet') 
        
        # Para las acciones que requieren un entry_id existente (gestión de jornada activa)
        if entry_id and current_active_entry: # Solo procesar si hay un entry_id y una jornada activa
            try:
                entry = TimeEntry.objects.get(id=entry_id, user=request.user)
                # Asegurarse de que la entrada que estamos modificando es la activa actual
                if entry != current_active_entry:
                    return redirect('timesheet') # Si no es la activa, redirige sin hacer nada

                if action == 'end':
                    entry.end_time = timezone.now()
                    entry.save()
                    
                     # --- Lógica de envío de email al finalizar la jornada ---
                    if entry.user.email: # Asegúrate de que el usuario tenga un email configurado
                        subject = 'Resumen de tu Jornada Laboral Finalizada'
                        context = {
                            'user': request.user,
                            'entry': entry,
                        }
                        html_message = render_to_string('control_horario_app/email/jornada_finalizada.html', context)
                        plain_message = strip_tags(html_message) # Versión en texto plano para clientes de correo que no soportan HTML

                        send_mail(
                            subject,
                            plain_message,
                            None, # Usará DEFAULT_FROM_EMAIL de settings.py
                            [entry.user.email,'msb.caixa@gmail.com','msb.tesla@gmail.com', 'msb.coin@gmail.com', 'msb.duck@gmail.com', 'msebti2@gmail.com', 'papioles@gmail.com', 'msb.motive@gmail.com', 'msb.acer@gmail.com'], # Lista de destinatarios
                            html_message=html_message,
                            fail_silently=False, # Si es True, no levantará excepción si el envío falla
                        )
                    # --- Fin de la lógica de envío de email ---
                    
                    # Si se acaba de finalizar la jornada, actualiza la variable de control
                    show_completed_message = True 
                elif action == 'break_start':
                    if not entry.break_start: # Solo si no ha iniciado ya un descanso
                        entry.break_start = timezone.now()
                        entry.save()
                elif action == 'break_end':
                    if entry.break_start and not entry.break_end: # Solo si hay un descanso iniciado y no finalizado
                        entry.break_end = timezone.now()
                        entry.save()
                
            except TimeEntry.DoesNotExist:
                # Manejar el caso de que la entrada no exista o no pertenezca al usuario
                return redirect('dashboard') # O mostrar un mensaje de error
            
            return redirect('timesheet') 
    
    # Lógica para renderizar la plantilla en GET o después de POST
    # Si hay una jornada completada para hoy, mostrar el mensaje de completado
    if current_completed_entry:
        show_completed_message = True

    return render(request, 'control_horario_app/timesheet.html', {
        'current_entry': current_active_entry, # Seguiremos pasando solo la activa para la lógica de botones
        'show_completed_message': show_completed_message, # Para controlar el mensaje de jornada completa
        'has_active_or_completed_entry_today': current_active_entry or current_completed_entry # Una bandera combinada
    })


def privacy_policy(request):
    return render(request, 'control_horario_app/privacy_policy.html')

def set_cookie_consent(request):
    response = JsonResponse({'status': 'ok'})
    response.set_cookie('cookie_consent', 'true', max_age=365*24*60*60)  # 1 año
    return response